import sys
import openpyxl
from sqlalchemy import text
from app import create_app
from models import db, Product

def run_migration():
    app = create_app()
    with app.app_context():
        print("1. Adding 'brand' column to product table if not exists...", flush=True)
        try:
            db.session.execute(text("ALTER TABLE product ADD COLUMN IF NOT EXISTS brand VARCHAR(100)"))
            db.session.commit()
            print("Successfully added brand column (or it already existed).", flush=True)
        except Exception as e:
            db.session.rollback()
            print(f"Error adding brand column: {e}", flush=True)
            return

        print("2. Parsing Stock Summary Excel file...", flush=True)
        try:
            wb = openpyxl.load_workbook(r'c:\Users\sriram\Downloads\erp-prd\Stock Summary.xlsx', data_only=True)
            sheet = wb.active
        except Exception as e:
            print(f"Error opening Stock Summary.xlsx: {e}", flush=True)
            return

        current_group = "Default"
        excel_products = {} # name_lower -> (group/category, brand)

        for row_idx in range(1, sheet.max_row + 1):
            cell_a = sheet.cell(row=row_idx, column=1).value
            cell_b = sheet.cell(row=row_idx, column=2).value # Item Name
            cell_f = sheet.cell(row=row_idx, column=6).value # Category (which is brand)
            
            if cell_a and str(cell_a).strip().startswith("Group Name:"):
                current_group = str(cell_a).strip().replace("Group Name:", "").strip()
            elif cell_b and str(cell_b).strip() not in ["", "Item", "Code", "Unit", "HSN/SAC Code", "Category", "Tax Code Name", "Opening", "Closing"]:
                item_name = str(cell_b).strip()
                brand = str(cell_f).strip() if cell_f else "Default"
                if brand.lower() == "none":
                    brand = "Default"
                excel_products[item_name.lower()] = (current_group, brand)

        print(f"Parsed {len(excel_products)} products from Excel sheet.", flush=True)

        print("3. Querying products in database...", flush=True)
        db_products = Product.query.all()
        print(f"Found {len(db_products)} products in the database.", flush=True)

        # Group product names by their target (category, brand) to run bulk update queries
        grouped_updates = {} # (category, brand) -> list of names
        for p in db_products:
            p_name_lower = p.name.lower()
            if p_name_lower in excel_products:
                group, brand = excel_products[p_name_lower]
            else:
                brand = p.category or "Default"
                group = "Uncategorized"
                
            key = (group, brand)
            if key not in grouped_updates:
                grouped_updates[key] = []
            grouped_updates[key].append(p.name)

        print(f"4. Executing bulk updates for {len(grouped_updates)} category/brand combinations...", flush=True)
        for (group, brand), names in grouped_updates.items():
            print(f"Updating {len(names)} products to Category: '{group}', Brand: '{brand}'...", flush=True)
            for i in range(0, len(names), 100):
                batch_names = names[i:i+100]
                db.session.query(Product).filter(Product.name.in_(batch_names)).update(
                    {Product.category: group, Product.brand: brand},
                    synchronize_session=False
                )
                
        db.session.commit()
        print("Successfully migrated DB!", flush=True)

if __name__ == '__main__':
    run_migration()
