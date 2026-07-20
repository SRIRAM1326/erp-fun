from flask import Blueprint, request, jsonify
from werkzeug.security import generate_password_hash, check_password_hash
from flask_jwt_extended import create_access_token, jwt_required, get_jwt_identity, get_jwt
from models import db, User, QRCode, PointsTransaction, Reward, Campaign, Invoice, Redemption, Product, Configuration
import uuid
import random
from datetime import datetime
from sqlalchemy import func

api = Blueprint('api', __name__)

@api.route('/auth/register', methods=['POST'])
def register():
    data = request.get_json()
    if User.query.filter_by(email=data.get('email')).first():
        return jsonify({'message': 'User already exists'}), 400
        
    # Generate referral code
    biz_name = data.get('business_name') or data.get('name') or 'USER'
    ref_code = f"REF-{biz_name.replace(' ', '').upper()[:5]}-{random.randint(1000, 9999)}"
        
    hashed_password = generate_password_hash(data.get('password'))
    role = data.get('role', 'buyer')
    
    # Default verification states
    is_verified = True
    verification_status = 'verified'
    
    referrer_id = None
    referrer_code = data.get('referral_code_used')
    if referrer_code:
        referrer = User.query.filter_by(referral_code=referrer_code).first()
        if referrer:
            referrer_id = referrer.id
            if referrer.role == 'rep':
                # Buyers referred by representatives are set to pending verification
                is_verified = False
                verification_status = 'pending'
                
    if role == 'rep':
        is_verified = True
        verification_status = 'verified'
        
    new_user = User(
        name=data.get('name'),
        email=data.get('email'),
        password_hash=hashed_password,
        role=role,
        business_name=data.get('business_name'),
        phone=data.get('phone'),
        referral_code=ref_code,
        referrer_id=referrer_id,
        is_verified=is_verified,
        verification_status=verification_status
    )
            
    db.session.add(new_user)
    db.session.commit()
    return jsonify({'message': 'User created successfully', 'referral_code': ref_code}), 201

@api.route('/auth/login', methods=['POST'])
def login():
    data = request.get_json()
    user = User.query.filter_by(email=data.get('email')).first()
    
    if user and check_password_hash(user.password_hash, data.get('password')):
        access_token = create_access_token(
            identity=str(user.id), 
            additional_claims={'id': user.id, 'role': user.role}
        )
        return jsonify({
            'token': access_token,
            'user': {
                'id': user.id,
                'name': user.name,
                'role': user.role,
                'points': user.total_points
            }
        }), 200
        
    return jsonify({'message': 'Invalid credentials'}), 401

@api.route('/admin/qr/generate', methods=['POST'])
@jwt_required()
def generate_qr():
    current_user = get_jwt()
    if current_user['role'] != 'admin':
        return jsonify({'message': 'Unauthorized'}), 403
        
    data = request.get_json()
    points = data.get('points', 100)
    
    new_qr = QRCode(
        code_value=str(uuid.uuid4()),
        points=points
    )
    db.session.add(new_qr)
    db.session.commit()
    
    return jsonify({'message': 'QR code generated', 'code_value': new_qr.code_value}), 201

@api.route('/admin/buyers', methods=['GET'])
@jwt_required()
def get_buyers():
    current_user = get_jwt()
    if current_user['role'] != 'admin':
        return jsonify({'message': 'Unauthorized'}), 403
        
    buyers = User.query.filter_by(role='buyer').all()
    result = [{
        'id': b.id,
        'name': b.name,
        'email': b.email,
        'business_name': b.business_name,
        'customer_code': b.customer_code or '0',
        'phone': b.phone or '0',
        'address': b.address or '0',
        'city': b.city or '0',
        'state': b.state or '0',
        'total_points': b.total_points,
        'tier': b.tier,
        'referral_code': b.referral_code,
        'is_verified': b.is_verified,
        'verification_status': b.verification_status
    } for b in buyers]
    
    return jsonify(result), 200

@api.route('/buyer/scan', methods=['POST'])
@jwt_required()
def scan_qr():
    current_user = get_jwt()
    if current_user['role'] != 'buyer':
        return jsonify({'message': 'Only buyers can scan QR codes'}), 403
        
    data = request.get_json()
    code_value = data.get('code_value')
    
    qr = QRCode.query.filter_by(code_value=code_value).first()
    if not qr:
        return jsonify({'message': 'Invalid QR code'}), 404
        
    if qr.status != 'active':
        return jsonify({'message': 'QR code already scanned or expired'}), 400
        
    buyer = User.query.get(current_user['id'])
    
    # 1. Dynamic Campaigns Check
    now = datetime.utcnow()
    active_campaign = Campaign.query.filter(
        Campaign.status == 'active',
        Campaign.start_date <= now,
        Campaign.end_date >= now
    ).first()
    
    earned_points = qr.points
    message = f'Success! {earned_points} points added.'
    
    if active_campaign:
        earned_points = int(earned_points * active_campaign.multiplier)
        message = f'Campaign Active ({active_campaign.name})! {earned_points} points added.'
    
    # Update QR
    qr.status = 'scanned'
    qr.scanned_at = now
    qr.buyer_id = buyer.id
    buyer.total_points += earned_points
    
    db.session.add(PointsTransaction(
        buyer_id=buyer.id, points=earned_points, transaction_type='credit', source='scan'
    ))
    
    # 2. Referral Bonus Check (Check if this is their first ever scan)
    has_scanned_before = PointsTransaction.query.filter_by(buyer_id=buyer.id, source='scan').count() > 1
    if not has_scanned_before and buyer.referrer_id:
        # Issue bonus
        referrer = User.query.get(buyer.referrer_id)
        if referrer:
            bonus = 5000
            buyer.total_points += bonus
            referrer.total_points += bonus
            
            db.session.add(PointsTransaction(
                buyer_id=buyer.id, points=bonus, transaction_type='credit', source='referral_bonus'
            ))
            db.session.add(PointsTransaction(
                buyer_id=referrer.id, points=bonus, transaction_type='credit', source='referral_bonus'
            ))
            message += f' Plus 5000 Referral Bonus!'

    db.session.commit()
    
    return jsonify({
        'message': message,
        'total_points': buyer.total_points
    }), 200

@api.route('/buyer/dashboard', methods=['GET'])
@jwt_required()
def buyer_dashboard():
    current_user = get_jwt()
    if current_user['role'] != 'buyer':
        return jsonify({'message': 'Unauthorized'}), 403
        
    buyer = User.query.get(current_user['id'])
    recent_txns = PointsTransaction.query.filter_by(buyer_id=buyer.id).order_by(PointsTransaction.created_at.desc()).limit(5).all()
    
    txns = [{
        'id': t.id,
        'points': t.points,
        'type': t.transaction_type,
        'source': t.source,
        'date': t.created_at.isoformat()
    } for t in recent_txns]
    
    return jsonify({
        'points': buyer.total_points,
        'tier': buyer.tier,
        'referral_code': buyer.referral_code,
        'recent_transactions': txns
    }), 200

@api.route('/buyer/history', methods=['GET'])
@jwt_required()
def buyer_history():
    current_user = get_jwt()
    if current_user['role'] != 'buyer':
        return jsonify({'message': 'Unauthorized'}), 403
        
    buyer = User.query.get(current_user['id'])
    txns = PointsTransaction.query.filter_by(buyer_id=buyer.id).order_by(PointsTransaction.created_at.desc()).all()
    
    history = [{
        'id': t.id,
        'points': t.points,
        'type': t.transaction_type,
        'source': t.source,
        'date': t.created_at.isoformat()
    } for t in txns]
    
    return jsonify(history), 200

@api.route('/buyer/leaderboard', methods=['GET'])
@jwt_required()
def get_leaderboard():
    # Get top 10 buyers by points
    top_buyers = User.query.filter_by(role='buyer').order_by(User.total_points.desc()).limit(10).all()
    
    leaderboard = []
    for i, b in enumerate(top_buyers):
        # Anonymize slightly (e.g. "Apex Distributors" -> "A. Distributors")
        parts = b.business_name.split(' ') if b.business_name else b.name.split(' ')
        display_name = f"{parts[0][0]}. {' '.join(parts[1:])}" if len(parts) > 1 else b.business_name or b.name
        
        leaderboard.append({
            'rank': i + 1,
            'name': display_name,
            'points': b.total_points,
            'tier': b.tier
        })
        
    return jsonify(leaderboard), 200

@api.route('/admin/reports/liability', methods=['GET'])
@jwt_required()
def get_liability():
    current_user = get_jwt()
    if current_user['role'] != 'admin':
        return jsonify({'message': 'Unauthorized'}), 403
        
    total_points = db.session.query(func.sum(User.total_points)).filter(User.role == 'buyer').scalar() or 0
    # Assume 1 point = 0.10 currency units (e.g. Rupees)
    financial_liability = total_points * 0.10
    
    return jsonify({
        'total_unredeemed_points': total_points,
        'financial_liability': financial_liability,
        'currency': 'INR'
    }), 200

@api.route('/admin/reports/chart-data', methods=['GET'])
@jwt_required()
def get_chart_data():
    current_user = get_jwt()
    if current_user['role'] != 'admin':
        return jsonify({'message': 'Unauthorized'}), 403
        
    # Return dummy aggregated data for the Recharts implementation to work immediately
    data = [
        {"name": "Jan", "issued": 4000, "redeemed": 2400},
        {"name": "Feb", "issued": 3000, "redeemed": 1398},
        {"name": "Mar", "issued": 2000, "redeemed": 9800},
        {"name": "Apr", "issued": 2780, "redeemed": 3908},
        {"name": "May", "issued": 1890, "redeemed": 4800},
        {"name": "Jun", "issued": 2390, "redeemed": 3800},
        {"name": "Jul", "issued": 3490, "redeemed": 4300},
    ]
    return jsonify(data), 200

@api.route('/admin/reports/analytics', methods=['GET'])
@jwt_required()
def get_analytics_report():
    current_user = get_jwt()
    if current_user['role'] != 'admin':
        return jsonify({'message': 'Unauthorized'}), 403

    # Liability
    total_points = db.session.query(func.sum(User.total_points)).filter(User.role == 'buyer').scalar() or 0
    financial_liability = total_points * 0.10

    # Chart data
    chart_data = [
        {"name": "2026-03", "issued": 4000, "redeemed": 1200},
        {"name": "2026-04", "issued": 6500, "redeemed": 2500},
        {"name": "2026-05", "issued": 8000, "redeemed": 3200},
        {"name": "2026-06", "issued": 10500, "redeemed": 4100},
        {"name": "2026-07", "issued": 14000, "redeemed": 5800},
    ]

    # Top buyers
    top_buyers_db = User.query.filter_by(role='buyer').order_by(User.total_points.desc()).limit(5).all()
    top_buyers = [{
        'id': b.id,
        'name': b.business_name or b.name,
        'total_points': b.total_points,
        'tier': b.tier or 'Silver'
    } for b in top_buyers_db]

    # Approaching customers (monthly spend threshold)
    latest_config = Configuration.query.order_by(Configuration.version.desc()).first()
    threshold = latest_config.high_spend_threshold if latest_config else 200000.0

    buyers = User.query.filter_by(role='buyer').limit(10).all()
    approaching_customers = []
    for b in buyers:
        now = datetime.utcnow()
        month_start = datetime(now.year, now.month, 1)
        monthly_spend = db.session.query(func.sum(Invoice.amount)).filter(
            Invoice.buyer_id == b.id,
            Invoice.status == 'paid',
            Invoice.created_at >= month_start
        ).scalar() or 0.0

        percentage = min(100.0, (monthly_spend / threshold * 100.0)) if threshold > 0 else 0.0
        approaching_customers.append({
            'id': b.id,
            'name': b.business_name or b.name,
            'email': b.email,
            'monthly_spend': monthly_spend,
            'threshold': threshold,
            'percentage': round(percentage, 1)
        })

    # Payout summaries (reps)
    reps = User.query.filter_by(role='rep').all()
    payout_summaries = []
    current_month_str = datetime.utcnow().strftime('%Y-%m')
    for r in reps:
        paid_invoices = Invoice.query.filter_by(rep_id=r.id, status='paid').all()
        credited_points = sum((inv.points_rep or 0) for inv in paid_invoices)
        if credited_points > 0:
            payout_summaries.append({
                'rep_name': r.name,
                'month': current_month_str,
                'commission_points': credited_points,
                'amount_rupees': credited_points
            })

    return jsonify({
        'liability': {
            'total_unredeemed_points': total_points,
            'financial_liability': financial_liability,
            'currency': 'INR'
        },
        'chart_data': chart_data,
        'top_buyers': top_buyers,
        'approaching_customers': approaching_customers,
        'payout_summaries': payout_summaries
    }), 200

@api.route('/admin/campaigns', methods=['GET'])
@jwt_required()
def get_campaigns():
    campaigns = Campaign.query.all()
    return jsonify([{
        'id': c.id,
        'name': c.name,
        'multiplier': c.multiplier,
        'start_date': c.start_date.isoformat(),
        'end_date': c.end_date.isoformat(),
        'status': c.status
    } for c in campaigns]), 200

@api.route('/rewards', methods=['GET'])
def get_rewards():
    rewards = Reward.query.filter_by(status='active').all()
    result = [{
        'id': r.id,
        'name': r.name,
        'description': r.description,
        'points_required': r.points_required
    } for r in rewards]
    return jsonify(result), 200

# ==================== REPRESENTATIVE PORTAL ROUTES ====================

@api.route('/rep/dashboard', methods=['GET'])
@jwt_required()
def rep_dashboard():
    current_user = get_jwt()
    if current_user['role'] != 'rep':
        return jsonify({'message': 'Unauthorized'}), 403
        
    rep = User.query.get(current_user['id'])
    
    # Query linked invoices
    invoices = Invoice.query.filter_by(rep_id=rep.id).all()
    
    total_invoices_count = len(invoices)
    total_value = sum(inv.amount for inv in invoices)
    
    # Rep earns 1% of invoice value. Credits only after invoice is paid.
    pending_points = sum(int(inv.amount * 0.01) for inv in invoices if inv.status == 'pending')
    credited_points = sum((inv.points_rep or 0) for inv in invoices if inv.status == 'paid')
    
    # Referred buyers
    buyers = User.query.filter_by(referrer_id=rep.id).all()
    buyer_list = [{
        'id': b.id,
        'name': b.name,
        'business_name': b.business_name,
        'email': b.email,
        'is_verified': b.is_verified,
        'status': b.verification_status,
        'date': b.created_at.isoformat()
    } for b in buyers]
    
    # Recent point transactions
    recent_txns = PointsTransaction.query.filter_by(buyer_id=rep.id).order_by(PointsTransaction.created_at.desc()).limit(5).all()
    txns = [{
        'id': t.id,
        'points': t.points,
        'type': t.transaction_type,
        'source': t.source,
        'date': t.created_at.isoformat()
    } for t in recent_txns]
    
    return jsonify({
        'points': rep.total_points,
        'tier': 'Gold' if rep.total_points > 5000 else 'Silver',
        'referral_code': rep.referral_code,
        'total_invoices': total_invoices_count,
        'total_value': total_value,
        'pending_points': pending_points,
        'credited_points': credited_points,
        'referred_buyers': buyer_list,
        'recent_transactions': txns
    }), 200

@api.route('/rep/invoices', methods=['GET'])
@jwt_required()
def rep_invoices():
    current_user = get_jwt()
    if current_user['role'] != 'rep':
        return jsonify({'message': 'Unauthorized'}), 403
        
    rep = User.query.get(current_user['id'])
    invoices = Invoice.query.filter_by(rep_id=rep.id).order_by(Invoice.created_at.desc()).all()
    
    result = []
    for inv in invoices:
        buyer = User.query.get(inv.buyer_id)
        result.append({
            'id': inv.id,
            'invoice_number': inv.invoice_number,
            'buyer_name': buyer.business_name or buyer.name if buyer else 'Unknown',
            'amount': inv.amount,
            'status': inv.status,
            'created_at': inv.created_at.isoformat(),
            'paid_at': inv.paid_at.isoformat() if inv.paid_at else None,
            'points_customer': inv.points_customer,
            'points_rep': inv.points_rep
        })
        
    return jsonify(result), 200

@api.route('/rep/invoices/link', methods=['POST'])
@jwt_required()
def rep_link_invoice():
    current_user = get_jwt()
    if current_user['role'] != 'rep':
        return jsonify({'message': 'Unauthorized'}), 403
        
    rep = User.query.get(current_user['id'])
    data = request.get_json()
    invoice_number = data.get('invoice_number', '').strip().upper()
    
    if not invoice_number:
        return jsonify({'message': 'Invoice number is required'}), 400
        
    invoice = Invoice.query.filter_by(invoice_number=invoice_number).first()
    
    if not invoice:
        # For seamless demo testing, auto-create a mock invoice for their referred buyer or default buyer
        buyer = User.query.filter_by(referrer_id=rep.id).first() or User.query.filter_by(role='buyer').first()
        if not buyer:
            return jsonify({'message': 'No buyer available to create invoice for'}), 400
            
        # Create a mock pending invoice
        import random
        invoice = Invoice(
            invoice_number=invoice_number,
            buyer_id=buyer.id,
            rep_id=rep.id,
            amount=float(random.randint(15, 180) * 1000), # e.g. ₹15,000 to ₹1,80,000
            status='pending',
            created_at=datetime.utcnow()
        )
        db.session.add(invoice)
        db.session.commit()
        return jsonify({
            'message': f'Mock invoice {invoice_number} created and linked successfully.',
            'invoice': {
                'id': invoice.id,
                'invoice_number': invoice.invoice_number,
                'amount': invoice.amount,
                'status': invoice.status
            }
        }), 201
        
    # Link existing invoice
    if invoice.rep_id == rep.id:
        return jsonify({'message': 'Invoice is already linked to you'}), 400
        
    invoice.rep_id = rep.id
    db.session.commit()
    
    return jsonify({'message': 'Invoice linked successfully'}), 200

@api.route('/rep/redeem', methods=['POST'])
@jwt_required()
def rep_redeem():
    current_user = get_jwt()
    if current_user['role'] != 'rep':
        return jsonify({'message': 'Unauthorized'}), 403
        
    rep = User.query.get(current_user['id'])
    data = request.get_json()
    
    points = int(data.get('points', 0))
    redemption_type = data.get('redemption_type') # 'cash' or 'product'
    details = data.get('details') # bank transfer details or product catalog selection
    
    if points <= 0:
        return jsonify({'message': 'Invalid points value'}), 400
        
    if rep.total_points < points:
        return jsonify({'message': 'Insufficient points balance'}), 400
        
    # Deduct points immediately
    rep.total_points -= points
    db.session.add(PointsTransaction(
        buyer_id=rep.id,
        points=points,
        transaction_type='debit',
        source='redemption',
        created_at=datetime.utcnow()
    ))
    
    # Create redemption request
    red = Redemption(
        user_id=rep.id,
        points_redeemed=points,
        redemption_type=redemption_type,
        details=details,
        status='pending',
        created_at=datetime.utcnow()
    )
    db.session.add(red)
    db.session.commit()
    
    return jsonify({
        'message': 'Redemption request submitted successfully.',
        'points_balance': rep.total_points
    }), 200

@api.route('/rep/history', methods=['GET'])
@jwt_required()
def rep_history():
    current_user = get_jwt()
    if current_user['role'] != 'rep':
        return jsonify({'message': 'Unauthorized'}), 403
        
    rep = User.query.get(current_user['id'])
    txns = PointsTransaction.query.filter_by(buyer_id=rep.id).order_by(PointsTransaction.created_at.desc()).all()
    
    result = [{
        'id': t.id,
        'points': t.points,
        'type': t.transaction_type,
        'source': t.source,
        'date': t.created_at.isoformat()
    } for t in txns]
    
    return jsonify(result), 200


# ==================== ADMIN PORTAL ROUTES FOR MARKETING ====================

@api.route('/admin/invoices', methods=['GET'])
@jwt_required()
def admin_invoices():
    current_user = get_jwt()
    if current_user['role'] != 'admin':
        return jsonify({'message': 'Unauthorized'}), 403
        
    invoices = Invoice.query.order_by(Invoice.created_at.desc()).all()
    result = []
    for inv in invoices:
        buyer = User.query.get(inv.buyer_id)
        rep = User.query.get(inv.rep_id) if inv.rep_id else None
        result.append({
            'id': inv.id,
            'invoice_number': inv.invoice_number,
            'buyer_id': inv.buyer_id,
            'buyer_name': buyer.business_name or buyer.name if buyer else 'N/A',
            'rep_id': inv.rep_id,
            'rep_name': rep.name if rep else 'N/A',
            'amount': inv.amount,
            'status': inv.status,
            'created_at': inv.created_at.isoformat(),
            'paid_at': inv.paid_at.isoformat() if inv.paid_at else None,
            'points_customer': inv.points_customer,
            'points_rep': inv.points_rep
        })
    return jsonify(result), 200

@api.route('/admin/invoices', methods=['POST'])
@jwt_required()
def admin_create_invoice():
    current_user = get_jwt()
    if current_user['role'] != 'admin':
        return jsonify({'message': 'Unauthorized'}), 403
        
    data = request.get_json()
    invoice_number = data.get('invoice_number', '').strip().upper()
    buyer_id = data.get('buyer_id')
    amount = float(data.get('amount', 0))
    rep_id = data.get('rep_id')
    products_str = data.get('products', '').strip()
    
    if not invoice_number or not buyer_id or amount <= 0:
        return jsonify({'message': 'Invoice number, buyer ID and amount are required'}), 400
        
    if Invoice.query.filter_by(invoice_number=invoice_number).first():
        return jsonify({'message': 'Invoice number already exists'}), 400
        
    latest_config = Configuration.query.order_by(Configuration.version.desc()).first()
    config_version = latest_config.version if latest_config else 1
        
    inv = Invoice(
        invoice_number=invoice_number,
        buyer_id=buyer_id,
        rep_id=rep_id,
        amount=amount,
        status='pending',
        products=products_str,
        config_version=config_version,
        created_at=datetime.utcnow()
    )
    db.session.add(inv)
    db.session.commit()
    
    return jsonify({'message': 'Invoice created successfully', 'id': inv.id}), 201

@api.route('/admin/invoices/<int:id>/pay', methods=['POST'])
@jwt_required()
def admin_pay_invoice(id):
    current_user = get_jwt()
    if current_user['role'] != 'admin':
        return jsonify({'message': 'Unauthorized'}), 403
        
    inv = Invoice.query.get(id)
    if not inv:
        return jsonify({'message': 'Invoice not found'}), 404
        
    if inv.status == 'paid':
        return jsonify({'message': 'Invoice is already paid'}), 400
        
    inv.status = 'paid'
    inv.paid_at = datetime.utcnow()
    
    # Load config active when the invoice was created
    config = Configuration.query.filter_by(version=inv.config_version).first()
    if not config:
        config = Configuration.query.order_by(Configuration.version.desc()).first()
        
    buyer_points, rep_points, rules_applied = calculate_invoice_points(inv, config)
        
    buyer = User.query.get(inv.buyer_id)
    if buyer:
        buyer.total_points += buyer_points
        inv.points_customer = buyer_points
        db.session.add(PointsTransaction(
            buyer_id=buyer.id,
            points=buyer_points,
            transaction_type='credit',
            source='invoice_payment',
            invoice_number=inv.invoice_number,
            rule_applied=rules_applied[:100],
            created_at=datetime.utcnow()
        ))
        
        # Check and award Loyalty Tier streak bonus if consecutive monthly purchase criteria is met
        check_and_award_loyalty_streak_bonus(buyer, config)
        
    # Representative Commission
    if rep_points > 0 and inv.rep_id:
        rep = User.query.get(inv.rep_id)
        if rep and rep.role == 'rep':
            rep.total_points += rep_points
            inv.points_rep = rep_points
            db.session.add(PointsTransaction(
                buyer_id=rep.id,
                points=rep_points,
                transaction_type='credit',
                source='sales_commission',
                invoice_number=inv.invoice_number,
                rule_applied=rules_applied[:100],
                created_at=datetime.utcnow()
            ))
            
    db.session.commit()
    
    return jsonify({
        'message': 'Invoice paid successfully. Points credited.',
        'points_customer': buyer_points,
        'points_rep': rep_points
    }), 200

@api.route('/admin/buyers/<int:buyer_id>/verify', methods=['POST'])
@jwt_required()
def admin_verify_buyer(buyer_id):
    current_user = get_jwt()
    if current_user['role'] != 'admin':
        return jsonify({'message': 'Unauthorized'}), 403
        
    buyer = User.query.get(buyer_id)
    if not buyer:
        return jsonify({'message': 'Buyer not found'}), 404
        
    if buyer.is_verified:
        return jsonify({'message': 'Buyer is already verified'}), 400
        
    buyer.is_verified = True
    buyer.verification_status = 'verified'
    
    message = 'Buyer verified successfully.'
    
    # Trigger representative onboarding bonus if referred by a rep
    if buyer.referrer_id:
        referrer = User.query.get(buyer.referrer_id)
        if referrer and referrer.role == 'rep':
            # Award a one-time onboarding reward of 1000 points
            bonus_points = 1000
            referrer.total_points += bonus_points
            db.session.add(PointsTransaction(
                buyer_id=referrer.id,
                points=bonus_points,
                transaction_type='credit',
                source='onboarding_bonus',
                created_at=datetime.utcnow()
            ))
            message += f' Awarded +1,000 onboarding bonus to Representative {referrer.name}.'
            
    db.session.commit()
    return jsonify({'message': message}), 200

@api.route('/admin/redemptions', methods=['GET'])
@jwt_required()
def admin_redemptions():
    current_user = get_jwt()
    if current_user['role'] != 'admin':
        return jsonify({'message': 'Unauthorized'}), 403
        
    redemptions = Redemption.query.order_by(Redemption.created_at.desc()).all()
    result = []
    for r in redemptions:
        user = User.query.get(r.user_id)
        result.append({
            'id': r.id,
            'user_id': r.user_id,
            'user_name': user.name if user else 'N/A',
            'user_role': user.role if user else 'N/A',
            'points_redeemed': r.points_redeemed,
            'redemption_type': r.redemption_type,
            'details': r.details,
            'status': r.status,
            'created_at': r.created_at.isoformat()
        })
        
    return jsonify(result), 200

@api.route('/admin/redemptions/<int:id>/approve', methods=['POST'])
@jwt_required()
def admin_approve_redemption(id):
    current_user = get_jwt()
    if current_user['role'] != 'admin':
        return jsonify({'message': 'Unauthorized'}), 403
        
    red = Redemption.query.get(id)
    if not red:
        return jsonify({'message': 'Redemption request not found'}), 404
        
    if red.status != 'pending':
        return jsonify({'message': 'Redemption is already processed'}), 400
        
    red.status = 'approved'
    db.session.commit()
    
    return jsonify({'message': 'Redemption approved successfully'}), 200

@api.route('/admin/redemptions/<int:id>/reject', methods=['POST'])
@jwt_required()
def admin_reject_redemption(id):
    current_user = get_jwt()
    if current_user['role'] != 'admin':
        return jsonify({'message': 'Unauthorized'}), 403
        
    red = Redemption.query.get(id)
    if not red:
        return jsonify({'message': 'Redemption request not found'}), 404
        
    if red.status != 'pending':
        return jsonify({'message': 'Redemption is already processed'}), 400
        
    red.status = 'rejected'
    
    # Refund points
    user = User.query.get(red.user_id)
    if user:
        user.total_points += red.points_redeemed
        db.session.add(PointsTransaction(
            buyer_id=user.id,
            points=red.points_redeemed,
            transaction_type='credit',
            source='redemption_refund',
            created_at=datetime.utcnow()
        ))
        
    db.session.commit()
    return jsonify({'message': 'Redemption request rejected and points refunded'}), 200

@api.route('/admin/reps', methods=['GET'])
@jwt_required()
def admin_reps():
    current_user = get_jwt()
    if current_user['role'] != 'admin':
        return jsonify({'message': 'Unauthorized'}), 403

    reps = User.query.filter_by(role='rep').all()
    result = []
    latest_config = Configuration.query.order_by(Configuration.version.desc()).first()
    min_purchase = latest_config.referral_min_value if latest_config else 50000.0

    for r in reps:
        referred_buyers = User.query.filter_by(referrer_id=r.id, role='buyer').all()
        referred_buyers_count = len(referred_buyers)
        verified_buyers_count = sum(1 for b in referred_buyers if b.is_verified)

        invoices = Invoice.query.filter_by(rep_id=r.id).all()
        total_invoices_count = len(invoices)

        qualifying_invoices = [inv for inv in invoices if inv.amount >= min_purchase]
        qualifying_invoices_count = len(qualifying_invoices)

        paid_invoices = [inv for inv in invoices if inv.status == 'paid']
        pending_invoices = [inv for inv in invoices if inv.status == 'pending']

        credited_points = sum((inv.points_rep or 0) for inv in paid_invoices)
        pending_points = sum(int(inv.amount * (latest_config.referral_rate if latest_config else 0.01))
                             for inv in pending_invoices if inv.amount >= min_purchase)

        total_sales = sum(inv.amount for inv in invoices)
        new_shops = referred_buyers_count

        # Conversion rate = paid invoices / total invoices
        conversion_rate = round((len(paid_invoices) / total_invoices_count * 100), 1) if total_invoices_count > 0 else 0.0

        result.append({
            'id': r.id,
            'name': r.name,
            'email': r.email,
            'phone': r.phone,
            'referral_code': r.referral_code,
            'is_active': r.is_verified,
            'points_balance': r.total_points,
            'referred_buyers_count': referred_buyers_count,
            'verified_buyers_count': verified_buyers_count,
            'total_invoices_count': total_invoices_count,
            'qualifying_invoices_count': qualifying_invoices_count,
            'pending_points': pending_points,
            'credited_points': credited_points,
            'total_sales': total_sales,
            'new_shops': new_shops,
            'conversion_rate': conversion_rate,
            'total_invoices_value': total_sales,
            'total_commission_earned': credited_points,
        })

    return jsonify(result), 200

@api.route('/admin/reps', methods=['POST'])
@jwt_required()
def admin_create_rep():
    current_user = get_jwt()
    if current_user['role'] != 'admin':
        return jsonify({'message': 'Unauthorized'}), 403

    data = request.get_json()
    name = data.get('name', '').strip()
    email = data.get('email', '').strip()
    phone = data.get('phone', '').strip()
    password = data.get('password', 'rep123')

    if not name or not email:
        return jsonify({'message': 'Name and email are required'}), 400

    if User.query.filter_by(email=email).first():
        return jsonify({'message': 'Email already in use'}), 400

    ref_code = f"REF-{name.replace(' ', '').upper()[:6]}-{random.randint(1000, 9999)}"
    rep = User(
        name=name,
        email=email,
        phone=phone,
        password_hash=generate_password_hash(password),
        role='rep',
        referral_code=ref_code,
        is_verified=True,
        verification_status='verified'
    )
    db.session.add(rep)
    db.session.commit()
    return jsonify({'message': 'Representative created successfully', 'id': rep.id, 'referral_code': ref_code}), 201

@api.route('/admin/reps/<int:rep_id>', methods=['PATCH'])
@jwt_required()
def admin_update_rep(rep_id):
    current_user = get_jwt()
    if current_user['role'] != 'admin':
        return jsonify({'message': 'Unauthorized'}), 403

    rep = User.query.filter_by(id=rep_id, role='rep').first()
    if not rep:
        return jsonify({'message': 'Representative not found'}), 404

    data = request.get_json()
    if 'name' in data:
        rep.name = data['name'].strip()
    if 'email' in data:
        existing = User.query.filter_by(email=data['email']).first()
        if existing and existing.id != rep_id:
            return jsonify({'message': 'Email already in use'}), 400
        rep.email = data['email'].strip()
    if 'phone' in data:
        rep.phone = data['phone'].strip()
    if 'is_active' in data:
        rep.is_verified = bool(data['is_active'])
        rep.verification_status = 'verified' if data['is_active'] else 'deactivated'

    db.session.commit()
    return jsonify({'message': 'Representative updated successfully'}), 200

@api.route('/admin/reps/<int:rep_id>', methods=['DELETE'])
@jwt_required()
def admin_delete_rep(rep_id):
    current_user = get_jwt()
    if current_user['role'] != 'admin':
        return jsonify({'message': 'Unauthorized'}), 403

    rep = User.query.filter_by(id=rep_id, role='rep').first()
    if not rep:
        return jsonify({'message': 'Representative not found'}), 404

    db.session.delete(rep)
    db.session.commit()
    return jsonify({'message': 'Representative deleted successfully'}), 200

# ==================== EXCEL/CSV IMPORT ROUTES ====================


import openpyxl
import io

def excel_to_dicts(file_storage):
    """
    Reads an Excel file storage object using openpyxl, parses the first row as headers,
    and returns a list of dictionaries mapping headers to cell values.
    """
    file_bytes = file_storage.read()
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    
    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    dicts = []
    for row in rows[1:]:
        if all(cell is None for cell in row):
            continue
        row_dict = {}
        for header, cell in zip(headers, row):
            if header:
                row_dict[header] = cell
        dicts.append(row_dict)
    return dicts

def read_uploaded_file_rows(file):
    """
    Reads an uploaded file (.xlsx, .xls, .xlsm, .csv) into a list of tuples/lists.
    Supports fallback to CSV text parsing if openpyxl fails.
    """
    filename = (file.filename or '').lower()
    file_bytes = file.read()
    
    if filename.endswith('.csv'):
        import csv
        text_content = file_bytes.decode('utf-8', errors='ignore')
        reader = csv.reader(text_content.splitlines())
        return [list(row) for row in reader if any(str(c).strip() for c in row)]
    else:
        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            sheet = wb.active
            return list(sheet.iter_rows(values_only=True))
        except Exception:
            import csv
            text_content = file_bytes.decode('utf-8', errors='ignore')
            reader = csv.reader(text_content.splitlines())
            rows = [list(row) for row in reader if any(str(c).strip() for c in row)]
            if len(rows) > 0:
                return rows
            raise Exception("Unable to parse file. Please provide a valid Excel (.xlsx, .xls) or CSV file.")

@api.route('/admin/upload/products', methods=['POST'])
@jwt_required()
def admin_upload_products():
    current_user = get_jwt()
    if current_user['role'] != 'admin':
        return jsonify({'message': 'Unauthorized'}), 403
        
    if 'file' not in request.files:
        return jsonify({'message': 'No file uploaded'}), 400
        
    file = request.files['file']
    filename = (file.filename or '').lower()
    if not filename.endswith(('.xlsx', '.xls', '.xlsm', '.csv')):
        return jsonify({'message': 'Only Excel files (.xlsx, .xls) or CSV files are supported'}), 400
        
    try:
        excel_rows = read_uploaded_file_rows(file)
        if not excel_rows or len(excel_rows) < 1:
            return jsonify({'message': 'Excel sheet is empty or has no data rows'}), 400

        header_keywords = ['product', 'item', 'code', 'name', 'tag', 'bonus', 'rate', 'category', 'unit', 'brand', 'price']
        best_header_idx = 0
        max_matches = -1
        
        for idx, row in enumerate(excel_rows[:25]):
            if not row or all(c is None or str(c).strip() == "" for c in row):
                continue
            row_str = ' '.join(str(c).lower().strip() for c in row if c is not None)
            matches = sum(1 for kw in header_keywords if kw in row_str)
            if matches > max_matches:
                max_matches = matches
                best_header_idx = idx
                
        header_row_idx = best_header_idx

        headers = [str(cell).strip() if cell is not None else "" for cell in excel_rows[header_row_idx]]

        def find_col_index(headers, keywords):
            for idx, header in enumerate(headers):
                h_clean = header.lower().replace(' ', '').replace('(', '').replace(')', '').replace('₹', '').replace('_', '')
                for keyword in keywords:
                    if keyword in h_clean:
                        return idx
            return None

        item_idx = find_col_index(headers, ['item', 'productname', 'products', 'product', 'name', 'code', 'description', 'title'])
        if item_idx is None:
            item_idx = 0

        tag_idx = find_col_index(headers, ['tag', 'type', 'status'])
        bonus_points_idx = find_col_index(headers, ['bonuspoints', 'bonus', 'points', 'reward'])
        category_idx = find_col_index(headers, ['category', 'group', 'brand'])
        sales_rate_idx = find_col_index(headers, ['salesrate', 'rate', 'price', 'amount'])

        existing_products = {p.name: p for p in Product.query.all()}
        count = 0
        current_group = "Default"

        for row in excel_rows[header_row_idx + 1:]:
            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue

            first_cell = str(row[0] or "").strip()
            if first_cell.lower().startswith('group name:'):
                current_group = first_cell[len('group name:'):].strip()
                continue

            if item_idx >= len(row):
                continue
            item_val = row[item_idx]
            if item_val is None:
                continue
            name = str(item_val).strip()[:100]

            if not name or name.lower() in ['item', 'product', 'product name', 'product_name', 'name', 'total', 'grand total', 'subtotal'] or name.lower().startswith('group name:'):
                continue

            tag = 'normal'
            if tag_idx is not None and tag_idx < len(row) and row[tag_idx] is not None:
                tag_val = str(row[tag_idx]).strip().lower()
                if tag_val in ['normal', 'special', 'old_stock', 'double_points']:
                    tag = tag_val

            bonus_points = 0
            if bonus_points_idx is not None and bonus_points_idx < len(row) and row[bonus_points_idx] is not None:
                bp_val = row[bonus_points_idx]
                if isinstance(bp_val, int):
                    bonus_points = bp_val
                elif isinstance(bp_val, float):
                    bonus_points = int(bp_val)
                else:
                    try:
                        bonus_points = int(float(str(bp_val).replace(',', '').strip()))
                    except:
                        bonus_points = 0

            brand_val = "Default"
            if category_idx is not None and category_idx < len(row) and row[category_idx] is not None:
                brand_val = str(row[category_idx]).strip()[:100]
                if brand_val.lower() == 'none':
                    brand_val = 'Default'

            sales_rate = 0.0
            if sales_rate_idx is not None and sales_rate_idx < len(row) and row[sales_rate_idx] is not None:
                sr_val = row[sales_rate_idx]
                if isinstance(sr_val, (int, float)):
                    sales_rate = float(sr_val)
                else:
                    try:
                        sales_rate = float(str(sr_val).replace('₹', '').replace(',', '').strip())
                    except:
                        sales_rate = 0.0

            product = existing_products.get(name)
            if product:
                product.tag = tag
                product.bonus_points = bonus_points
                product.category = current_group
                product.brand = brand_val
                if sales_rate > 0:
                    product.sales_rate = sales_rate
            else:
                product = Product(
                    name=name, 
                    tag=tag, 
                    bonus_points=bonus_points,
                    category=current_group,
                    brand=brand_val,
                    sales_rate=sales_rate
                )
                db.session.add(product)
                existing_products[name] = product

            count += 1
            if count % 50 == 0:
                db.session.commit()

        db.session.commit()
        return jsonify({'message': f'Successfully imported {count} products'}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'message': f'Failed to process product file: {str(e)}'}), 500

@api.route('/admin/upload/customers', methods=['POST'])
@jwt_required()
def admin_upload_customers():
    current_user = get_jwt()
    if current_user['role'] != 'admin':
        return jsonify({'message': 'Unauthorized'}), 403
        
    if 'file' not in request.files:
        return jsonify({'message': 'No file uploaded'}), 400
        
    file = request.files['file']
    filename = (file.filename or '').lower()
    if not filename.endswith(('.xlsx', '.xls', '.xlsm', '.csv')):
        return jsonify({'message': 'Only Excel files (.xlsx, .xls) or CSV files are supported'}), 400
        
    try:
        excel_rows = read_uploaded_file_rows(file)
        if not excel_rows or len(excel_rows) < 1:
            return jsonify({'message': 'Excel sheet is empty or has no data rows'}), 400

        header_keywords = ['code', 'customer', 'name', 'address', 'city', 'state', 'phone', 'mail', 'email', 'buyer', 'party', 'mobile', 'contact']
        best_header_idx = 0
        max_matches = -1
        
        for idx, row in enumerate(excel_rows[:25]):
            if not row or all(c is None or str(c).strip() == "" for c in row):
                continue
            row_str = ' '.join(str(c).lower().strip() for c in row if c is not None)
            matches = sum(1 for kw in header_keywords if kw in row_str)
            if matches > max_matches:
                max_matches = matches
                best_header_idx = idx
                
        header_row_idx = best_header_idx

        headers = [str(cell).strip() if cell is not None else "" for cell in excel_rows[header_row_idx]]

        def find_col_index(headers, keywords):
            for idx, header in enumerate(headers):
                h_clean = header.lower().replace(' ', '').replace('_', '').replace('-', '')
                for keyword in keywords:
                    if keyword in h_clean:
                        return idx
            return None

        code_idx = find_col_index(headers, ['customercode', 'code', 'custcode', 'id', 'customerno', 'no'])
        name_idx = find_col_index(headers, ['customername', 'name', 'buyername', 'storename', 'businessname', 'party'])
        address_idx = find_col_index(headers, ['address', 'street', 'location', 'addr'])
        city_idx = find_col_index(headers, ['city', 'town', 'district'])
        state_idx = find_col_index(headers, ['state', 'province', 'region'])
        phone_idx = find_col_index(headers, ['phone', 'mobile', 'contact', 'phonenumber', 'tel', 'cell'])
        mail_idx = find_col_index(headers, ['mail', 'email', 'emailaddress', 'mailid'])

        total_records = 0
        imported_count = 0
        updated_count = 0
        skipped_count = 0
        errors = []

        used_emails = set(b.email.lower() for b in User.query.all() if b.email)
        existing_buyers_by_code = {b.customer_code: b for b in User.query.filter(User.customer_code != None).all() if b.customer_code}
        existing_buyers_by_email = {b.email.lower(): b for b in User.query.all() if b.email}
        existing_buyers_by_name = {b.name.lower(): b for b in User.query.all() if b.name}
        existing_referral_codes = set(u.referral_code for u in User.query.filter(User.referral_code != None).all() if u.referral_code)

        for row_num, row in enumerate(excel_rows[header_row_idx + 1:], start=header_row_idx + 2):
            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue

            total_records += 1

            def get_val(col_idx):
                if col_idx is None or col_idx >= len(row):
                    return "0"
                raw = row[col_idx]
                if raw is None:
                    return "0"
                val = str(raw).strip()
                return val if val != "" else "0"

            code_val = get_val(code_idx)[:100]
            name_val = get_val(name_idx)[:100]
            address_val = get_val(address_idx)[:255]
            city_val = get_val(city_idx)[:100]
            state_val = get_val(state_idx)[:100]
            phone_val = get_val(phone_idx)[:100]
            mail_val = get_val(mail_idx)[:120]

            if code_val == "0" and name_val == "0":
                skipped_count += 1
                errors.append(f"Row {row_num}: Both Customer Code and Customer Name are empty. Skipped.")
                continue

            if mail_val != "0" and '@' in mail_val:
                target_email = mail_val.lower().strip()[:120]
            else:
                clean_key = (code_val if code_val != "0" else name_val).lower().replace(' ', '').replace('/', '').replace('\\', '').replace('_', '')[:60]
                target_email = f"cust_{clean_key}@example.com"

            clean_name = (name_val if name_val != "0" else f"Customer {code_val}")[:100]

            buyer = None
            if code_val != "0" and code_val in existing_buyers_by_code:
                buyer = existing_buyers_by_code[code_val]
            elif target_email in existing_buyers_by_email:
                buyer = existing_buyers_by_email[target_email]
            elif clean_name.lower() in existing_buyers_by_name:
                buyer = existing_buyers_by_name[clean_name.lower()]

            final_email = target_email
            if final_email in used_emails and (not buyer or (buyer.email and buyer.email.lower() != final_email)):
                import uuid
                if '@' in final_email:
                    prefix, domain = final_email.split('@', 1)
                else:
                    prefix, domain = final_email, 'example.com'
                prefix = prefix[:60]
                final_email = f"{prefix}_{uuid.uuid4().hex[:4]}@{domain}"
                while final_email in used_emails:
                    final_email = f"{prefix}_{uuid.uuid4().hex[:4]}@{domain}"

            if buyer:
                old_email = buyer.email.lower() if buyer.email else None
                buyer.name = clean_name
                buyer.business_name = clean_name
                if code_val != "0":
                    buyer.customer_code = code_val
                buyer.address = address_val
                buyer.city = city_val
                buyer.state = state_val
                buyer.phone = phone_val
                if (mail_val != "0" and '@' in mail_val) or not buyer.email:
                    if old_email and old_email in used_emails and old_email != final_email:
                        used_emails.discard(old_email)
                        existing_buyers_by_email.pop(old_email, None)
                    buyer.email = final_email
                    used_emails.add(final_email)
                    existing_buyers_by_email[final_email] = buyer
                updated_count += 1
            else:
                import uuid
                ref_code = f"REF-CUST-{uuid.uuid4().hex[:8].upper()}"
                while ref_code in existing_referral_codes:
                    ref_code = f"REF-CUST-{uuid.uuid4().hex[:8].upper()}"
                existing_referral_codes.add(ref_code)

                buyer = User(
                    name=clean_name,
                    business_name=clean_name,
                    email=final_email,
                    customer_code=code_val if code_val != "0" else f"CUST-{uuid.uuid4().hex[:6].upper()}",
                    password_hash=generate_password_hash('buyer123'),
                    role='buyer',
                    address=address_val,
                    city=city_val,
                    state=state_val,
                    phone=phone_val,
                    referral_code=ref_code,
                    is_verified=True,
                    verification_status='verified'
                )
                db.session.add(buyer)
                used_emails.add(final_email)
                existing_buyers_by_email[final_email] = buyer
                imported_count += 1

            if buyer.customer_code:
                existing_buyers_by_code[buyer.customer_code] = buyer
            if buyer.name:
                existing_buyers_by_name[buyer.name.lower()] = buyer

        db.session.commit()

        return jsonify({
            'message': f'Customer Master import completed. {imported_count} imported, {updated_count} updated, {skipped_count} skipped.',
            'summary': {
                'total_records': total_records,
                'imported': imported_count,
                'updated': updated_count,
                'skipped': skipped_count,
                'errors': errors
            }
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'message': f'Failed to process customer file: {str(e)}'}), 500

@api.route('/admin/upload/invoices', methods=['POST'])
@jwt_required()
def admin_upload_invoices():
    current_user = get_jwt()
    if current_user['role'] != 'admin':
        return jsonify({'message': 'Unauthorized'}), 403
        
    if 'file' not in request.files:
        return jsonify({'message': 'No file uploaded'}), 400
        
    file = request.files['file']
    filename = (file.filename or '').lower()
    if not filename.endswith(('.xlsx', '.xls', '.xlsm', '.csv')):
        return jsonify({'message': 'Only Excel files (.xlsx, .xls) or CSV files are supported'}), 400
        
    try:
        excel_rows = read_uploaded_file_rows(file)
        if not excel_rows or len(excel_rows) < 1:
            return jsonify({'message': 'Excel sheet is empty or has no data rows'}), 400
            
        # 1. Smart Header Row Selection by scoring candidate rows
        header_keywords = [
            'invoice', 'invoiceno', 'bill', 'billno', 'vch', 'voucher', 'no', 'number',
            'customer', 'buyer', 'party', 'client', 'name', 'store',
            'amount', 'total', 'netamt', 'val', 'price', 'salesman', 'rep', 'salesperson', 'item', 'product'
        ]
        
        best_header_idx = 0
        max_matches = -1
        
        for idx, row in enumerate(excel_rows[:25]):
            if not row or all(c is None or str(c).strip() == "" for c in row):
                continue
            row_str = ' '.join(str(c).lower().strip() for c in row if c is not None)
            matches = sum(1 for kw in header_keywords if kw in row_str)
            if matches > max_matches:
                max_matches = matches
                best_header_idx = idx
                
        header_row_idx = best_header_idx
        headers = [str(cell).strip() if cell is not None else "" for cell in excel_rows[header_row_idx]]
        
        def get_clean_header(h):
            return str(h or '').lower().replace(' ', '').replace('(', '').replace(')', '').replace('₹', '').replace('_', '').replace('-', '').replace('.', '')

        def is_serial_no_col(h):
            c = get_clean_header(h)
            return c in ['sno', 'slno', 'srno', 'serialno', 'serialnumber', 'seqno', 'rowno', 'index', 'sl'] or c.startswith(('sno', 'slno', 'srno'))

        # 2. Precision Column Index Matching
        # Invoice Number Column:
        no_idx = None
        for idx, h in enumerate(headers):
            if is_serial_no_col(h):
                continue
            ch = get_clean_header(h)
            if ch in ['no', 'invoiceno', 'invoicenumber', 'vchno', 'voucherno', 'billno', 'billnumber', 'docno', 'docnumber', 'invno']:
                no_idx = idx
                break

        if no_idx is None:
            for idx, h in enumerate(headers):
                if is_serial_no_col(h):
                    continue
                ch = get_clean_header(h)
                if any(kw in ch for kw in ['invoiceno', 'invoicenumber', 'vchno', 'voucherno', 'billno', 'billnumber', 'docno', 'docnumber', 'invno', 'refno', 'transactionid', 'receiptno']):
                    no_idx = idx
                    break

        if no_idx is None:
            for idx, h in enumerate(headers):
                if is_serial_no_col(h):
                    continue
                ch = get_clean_header(h)
                if any(kw in ch for kw in ['invoice', 'vch', 'voucher', 'bill', 'doc', 'ref', 'txn']):
                    no_idx = idx
                    break

        if no_idx is None:
            for idx, h in enumerate(headers):
                if not is_serial_no_col(h):
                    no_idx = idx
                    break
            if no_idx is None:
                no_idx = 0

        # Customer / Party Column:
        customer_idx = None
        for idx, h in enumerate(headers):
            ch = get_clean_header(h)
            if ch in ['customer', 'customername', 'party', 'partyname', 'buyer', 'buyername', 'store', 'storename', 'businessname', 'account', 'accountname', 'particulars']:
                customer_idx = idx
                break

        if customer_idx is None:
            for idx, h in enumerate(headers):
                ch = get_clean_header(h)
                if any(kw in ch for kw in ['customer', 'buyer', 'party', 'store', 'client', 'account', 'particulars', 'billed', 'name']):
                    customer_idx = idx
                    break

        if customer_idx is None:
            customer_idx = 1 if len(headers) > 1 else 0

        # Amount Column:
        amount_idx = None
        for idx, h in enumerate(headers):
            ch = get_clean_header(h)
            if ch in ['billamount', 'netamount', 'invoiceamount', 'grandtotal', 'totalamount', 'billamt', 'netamt', 'invamt', 'amount', 'total']:
                amount_idx = idx
                break

        if amount_idx is None:
            for idx, h in enumerate(headers):
                ch = get_clean_header(h)
                if any(kw in ch for kw in ['billamount', 'netamount', 'invoiceamount', 'grandtotal', 'totalamount', 'taxableamount', 'taxablevalue', 'totalvalue', 'netvalue', 'invoicevalue', 'billvalue', 'invvalue', 'debitamount', 'totalamt', 'amount', 'net', 'total', 'val', 'price']):
                    amount_idx = idx
                    break

        if amount_idx is None:
            amount_idx = 2 if len(headers) > 2 else (1 if len(headers) > 1 else 0)

        # Salesman Column:
        salesman_idx = None
        for idx, h in enumerate(headers):
            ch = get_clean_header(h)
            if any(kw in ch for kw in ['salesman', 'salesperson', 'salesrep', 'executive', 'representative', 'repemail', 'salesexecutive', 'employee', 'agent', 'rep', 'se']):
                salesman_idx = idx
                break

        # Item / Product Column:
        item_idx = None
        for idx, h in enumerate(headers):
            if idx in [no_idx, customer_idx, amount_idx]:
                continue
            ch = get_clean_header(h)
            if ch in ['item', 'itemname', 'product', 'productname', 'description', 'particulars', 'goods', 'stockitem']:
                item_idx = idx
                break

        if item_idx is None:
            for idx, h in enumerate(headers):
                if idx in [no_idx, customer_idx, amount_idx]:
                    continue
                ch = get_clean_header(h)
                if any(kw in ch for kw in ['productname', 'itemname', 'products', 'product', 'items', 'item', 'description', 'particulars', 'goods', 'title', 'stockitem']):
                    item_idx = idx
                    break

        # Item Amount Column:
        item_amount_idx = None
        for idx, h in enumerate(headers):
            if idx == amount_idx:
                continue
            ch = get_clean_header(h)
            if 'amount' in ch or 'amt' in ch:
                item_amount_idx = idx
                break

        # Brand Column:
        brand_idx = None
        for idx, h in enumerate(headers):
            ch = get_clean_header(h)
            if 'brand' in ch or 'manufacturer' in ch or 'make' in ch:
                brand_idx = idx
                break

        # Helper for parsing numeric amount safely
        def parse_amount(val):
            if val is None:
                return 0.0
            if isinstance(val, (int, float)):
                return float(val)
            s = str(val).strip()
            if not s:
                return 0.0
            import re
            cleaned = re.sub(r'[^\d.-]', '', s.replace(',', '').replace('/-', ''))
            try:
                return float(cleaned) if cleaned else 0.0
            except:
                return 0.0

        invoices_dict = {}
        last_inv_no = ""
        last_customer = ""
        last_salesman = ""

        for row in excel_rows[header_row_idx + 1:]:
            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue
                
            inv_no_raw = row[no_idx] if no_idx < len(row) else None
            inv_no = str(inv_no_raw).strip().upper() if inv_no_raw is not None else ""
            
            # Fallback to Ref No if No cell is blank or 0
            if not inv_no or inv_no in ["0", "NONE", "NULL"]:
                ref_no_raw = row[2] if len(row) > 2 else None
                if ref_no_raw is not None and str(ref_no_raw).strip():
                    inv_no = str(ref_no_raw).strip().upper()

            if not inv_no and last_inv_no:
                inv_no = last_inv_no

            if not inv_no or inv_no in ['NO', 'INVOICE NUMBER', 'BILL NO', 'TOTAL', 'GRAND TOTAL', 'SUBTOTAL', 'SUMMARY', 'REPORT', 'S.NO', 'SL.NO', 'SR.NO'] or inv_no.startswith('BRANCH:'):
                continue

            last_inv_no = inv_no

            customer_raw = row[customer_idx] if customer_idx < len(row) else None
            customer = str(customer_raw).strip() if customer_raw is not None else ""
            if not customer and last_customer and inv_no == last_inv_no:
                customer = last_customer

            if customer and customer.lower() not in ['customer', 'buyer', 'party', 'total', 'grand total', 'subtotal', 'summary', 'report']:
                last_customer = customer

            salesman_raw = row[salesman_idx] if (salesman_idx is not None and salesman_idx < len(row)) else None
            salesman = str(salesman_raw).strip() if salesman_raw is not None else ""
            if not salesman and last_salesman and inv_no == last_inv_no:
                salesman = last_salesman

            if salesman:
                last_salesman = salesman

            amount = parse_amount(row[amount_idx]) if amount_idx < len(row) else 0.0
            item_amount = parse_amount(row[item_amount_idx]) if (item_amount_idx is not None and item_amount_idx < len(row)) else 0.0
            item = str(row[item_idx] or '').strip() if (item_idx is not None and item_idx < len(row)) else ""
            brand = str(row[brand_idx] or '').strip() if (brand_idx is not None and brand_idx < len(row)) else ""

            if inv_no not in invoices_dict:
                invoices_dict[inv_no] = {
                    'customer': customer,
                    'salesman': salesman,
                    'amount': amount,
                    'items': [],
                    'brands': {},
                    'items_total_amount': 0.0
                }

            if customer and not invoices_dict[inv_no]['customer']:
                invoices_dict[inv_no]['customer'] = customer
            if salesman and not invoices_dict[inv_no]['salesman']:
                invoices_dict[inv_no]['salesman'] = salesman
            if amount > invoices_dict[inv_no]['amount']:
                invoices_dict[inv_no]['amount'] = amount

            if item:
                invoices_dict[inv_no]['items'].append(item)
                invoices_dict[inv_no]['items_total_amount'] += item_amount
                if brand and brand.lower() not in ['brand', 'none', 'null', '0']:
                    invoices_dict[inv_no]['brands'][item] = brand

        existing_products = {p.name: p for p in Product.query.all()}

        # Build multi-index buyer lookup table
        all_buyers = User.query.filter_by(role='buyer').all()
        buyer_lookup = {}
        for b in all_buyers:
            if b.name:
                buyer_lookup[b.name.lower().strip()] = b
            if b.business_name:
                buyer_lookup[b.business_name.lower().strip()] = b
            if b.customer_code:
                buyer_lookup[b.customer_code.lower().strip()] = b
            if b.email:
                buyer_lookup[b.email.lower().strip()] = b

        # Build multi-index rep lookup table
        all_reps = User.query.filter_by(role='rep').all()
        rep_lookup = {}
        for r in all_reps:
            if r.name:
                rep_lookup[r.name.lower().strip()] = r
            if r.email:
                rep_lookup[r.email.lower().strip()] = r
            if r.phone:
                rep_lookup[r.phone.lower().strip()] = r

        existing_referral_codes = set(u.referral_code for u in User.query.filter(User.referral_code != None).all() if u.referral_code)
        
        count = 0
        for invoice_number, info in invoices_dict.items():
            customer = info['customer']
            salesman = info['salesman']
            amount = info['amount']
            items_list = info['items']
            items_total = info['items_total_amount']
            
            if amount <= 0:
                amount = items_total
                
            if not customer or amount <= 0:
                continue
                
            products_str = ", ".join(items_list)
            
            for item in items_list:
                item_brand = info['brands'].get(item, 'Default')
                product = existing_products.get(item)
                if not product:
                    product = Product(name=item, tag='normal', bonus_points=0, brand=item_brand)
                    db.session.add(product)
                    existing_products[item] = product
                elif item_brand != 'Default' and (not product.brand or product.brand == 'Default'):
                    product.brand = item_brand
            db.session.commit()
            
            cust_key = customer.lower().strip()
            buyer = buyer_lookup.get(cust_key)
            if not buyer:
                clean_name = customer.lower().replace(' ', '')
                email = f"{clean_name}@example.com"
                buyer = buyer_lookup.get(email)
                if not buyer:
                    import uuid
                    ref_code = f"REF-BUYER-{uuid.uuid4().hex[:8].upper()}"
                    while ref_code in existing_referral_codes:
                        ref_code = f"REF-BUYER-{uuid.uuid4().hex[:8].upper()}"
                    existing_referral_codes.add(ref_code)

                    buyer = User(
                        name=customer,
                        email=email,
                        password_hash=generate_password_hash('buyer123'),
                        role='buyer',
                        business_name=f"{customer} Store",
                        referral_code=ref_code,
                        is_verified=True,
                        verification_status='verified'
                    )
                    db.session.add(buyer)
                    db.session.commit()
                buyer_lookup[cust_key] = buyer
                buyer_lookup[buyer.email.lower()] = buyer
            
            rep_id = None
            if salesman:
                rep_key = salesman.lower().strip()
                rep = rep_lookup.get(rep_key)
                if not rep:
                    clean_rep_name = salesman.lower().replace(' ', '')
                    rep_email = f"{clean_rep_name}@sales.com"
                    rep = rep_lookup.get(rep_email)
                    if not rep:
                        import uuid
                        ref_code = f"REF-REP-{uuid.uuid4().hex[:8].upper()}"
                        while ref_code in existing_referral_codes:
                            ref_code = f"REF-REP-{uuid.uuid4().hex[:8].upper()}"
                        existing_referral_codes.add(ref_code)

                        rep = User(
                            name=salesman,
                            email=rep_email,
                            password_hash=generate_password_hash('rep123'),
                            role='rep',
                            referral_code=ref_code,
                            is_verified=True,
                            verification_status='verified'
                        )
                        db.session.add(rep)
                        db.session.commit()
                    rep_lookup[rep_key] = rep
                    rep_lookup[rep.email.lower()] = rep
                rep_id = rep.id
                
            status = 'paid'
            invoice = Invoice.query.filter_by(invoice_number=invoice_number).first()
            is_new_payment = False
            
            if invoice:
                if invoice.status != 'paid' and status == 'paid':
                    is_new_payment = True
                invoice.buyer_id = buyer.id
                invoice.rep_id = rep_id
                invoice.amount = amount
                invoice.status = status
                invoice.products = products_str
            else:
                if status == 'paid':
                    is_new_payment = True
                
                latest_config = Configuration.query.order_by(Configuration.version.desc()).first()
                config_version = latest_config.version if latest_config else 1
                
                invoice = Invoice(
                    invoice_number=invoice_number,
                    buyer_id=buyer.id,
                    rep_id=rep_id,
                    amount=amount,
                    status=status,
                    products=products_str,
                    config_version=config_version,
                    created_at=datetime.utcnow()
                )
                db.session.add(invoice)
                
            db.session.commit()
            
            if is_new_payment:
                invoice.paid_at = datetime.utcnow()
                config = Configuration.query.filter_by(version=invoice.config_version).first()
                if not config:
                    config = Configuration.query.order_by(Configuration.version.desc()).first()
                    
                buyer_points, rep_points, rules_applied = calculate_invoice_points(invoice, config)
                
                buyer.total_points += buyer_points
                invoice.points_customer = buyer_points
                db.session.add(PointsTransaction(
                    buyer_id=buyer.id,
                    points=buyer_points,
                    transaction_type='credit',
                    source='invoice_payment',
                    invoice_number=invoice.invoice_number,
                    rule_applied=rules_applied[:100],
                    created_at=datetime.utcnow()
                ))
                
                if rep_points > 0 and invoice.rep_id:
                    rep = User.query.get(invoice.rep_id)
                    if rep and rep.role == 'rep':
                        rep.total_points += rep_points
                        invoice.points_rep = rep_points
                        db.session.add(PointsTransaction(
                            buyer_id=rep.id,
                            points=rep_points,
                            transaction_type='credit',
                            source='sales_commission',
                            invoice_number=invoice.invoice_number,
                            rule_applied=rules_applied[:100],
                            created_at=datetime.utcnow()
                        ))
                        
                db.session.commit()
                check_and_award_loyalty_streak_bonus(buyer, config)
                
            count += 1
            
        if count == 0:
            headers_str = ", ".join([h for h in headers if h]) if headers else "None"
            return jsonify({
                'message': f'No valid invoice records could be matched (0 records imported). Detected headers: [{headers_str}]. Please ensure your file has columns for Invoice Number, Customer/Party Name, and Amount.'
            }), 400
            
        return jsonify({'message': f'Successfully imported {count} invoices and updated dashboards.'}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'message': f'Failed to process invoice file: {str(e)}'}), 500


# ==================== POINTS CALCULATION HELPER ====================

def calculate_invoice_points(inv, config):
    days_elapsed = (inv.paid_at - inv.created_at).days if inv.paid_at else 0
    if days_elapsed < 0:
        days_elapsed = 0
        
    inv_pct = getattr(config, 'invoice_reward_percentage', 0.50) if config else 0.50
    if inv_pct is None:
        inv_pct = 0.50
    base_buyer_points = int(inv.amount * inv_pct)
    multiplier = 1.0
    rules = []
    rules.append(f"BR-01 (Base conversion {inv_pct * 100:.1f}%)")
    
    credit_period = config.credit_period
    cutoff = config.forfeiture_cutoff
    
    if days_elapsed <= credit_period:
        multiplier *= 2.0
        rules.append(f"BR-02 (Paid within {credit_period}d credit period - 2x)")
    elif days_elapsed > cutoff:
        multiplier *= 0.0
        rules.append(f"BR-04 (Forfeited - paid after {cutoff}d - 0x)")
    else:
        # scale linearly
        factor = 2.0 - ((days_elapsed - credit_period) / float(cutoff - credit_period)) * 2.0
        multiplier *= max(0.0, factor)
        rules.append(f"BR-03 (Paid in {days_elapsed}d - scale factor {factor:.2f}x)")
        
    # Product tags check
    flat_bonuses = 0
    has_double_points_product = False
    
    if inv.products:
        product_names = [p.strip() for p in inv.products.split(',') if p.strip()]
        for prod_name in product_names:
            product = Product.query.filter_by(name=prod_name).first()
            if product:
                if product.tag == 'special':
                    val = product.bonus_points or config.special_bonus
                    flat_bonuses += val
                    rules.append(f"BR-10 ({prod_name} Special product bonus +{val})")
                elif product.tag == 'old_stock':
                    val = product.bonus_points or config.old_stock_bonus
                    flat_bonuses += val
                    rules.append(f"BR-11 ({prod_name} Old Stock product bonus +{val})")
                elif product.tag == 'double_points':
                    has_double_points_product = True
                    rules.append(f"BR-09 ({prod_name} Double points product - 2x)")
                    
    if has_double_points_product:
        multiplier *= 2.0
        
    buyer_points = int(base_buyer_points * max(0.0, multiplier)) + flat_bonuses
    
    # Representative Commission points
    rep_points = 0
    if inv.rep_id:
        if inv.amount >= config.referral_min_value:
            rep_points = int(inv.amount * config.referral_rate)
            rules.append(f"BR-12 (Qualified Referral commission {config.referral_rate * 100:.1f}%)")
        else:
            rules.append(f"BR-13 (Did not qualify - amount below {config.referral_min_value})")
            
    return buyer_points, rep_points, ", ".join(rules)

def check_and_award_loyalty_streak_bonus(buyer, config):
    """
    Checks if a buyer has purchased >= loyalty_min_monthly_purchase in 
    loyalty_consecutive_months consecutive calendar months. If qualified and 
    not already awarded for the period, credits loyalty_bonus points.
    """
    if not buyer or not config:
        return False, 0

    req_months = getattr(config, 'loyalty_consecutive_months', 3) or 3
    min_spend = getattr(config, 'loyalty_min_monthly_purchase', 200000.0) or 200000.0
    bonus_pts = getattr(config, 'loyalty_bonus', 10000) or 10000

    if req_months <= 0 or min_spend <= 0 or bonus_pts <= 0:
        return False, 0

    paid_invoices = Invoice.query.filter_by(buyer_id=buyer.id, status='paid').all()
    if not paid_invoices:
        return False, 0

    monthly_totals = {}
    for inv in paid_invoices:
        dt = inv.paid_at or inv.created_at
        if dt:
            month_key = dt.strftime('%Y-%m')
            monthly_totals[month_key] = monthly_totals.get(month_key, 0.0) + inv.amount

    qualified_months = sorted([m for m, total in monthly_totals.items() if total >= min_spend])
    if len(qualified_months) < req_months:
        return False, 0

    def month_to_num(m_str):
        y, m = map(int, m_str.split('-'))
        return y * 12 + m

    month_nums = [month_to_num(m) for m in qualified_months]
    current_streak = 1
    for i in range(len(month_nums) - 1, 0, -1):
        if month_nums[i] - month_nums[i - 1] == 1:
            current_streak += 1
        else:
            break

    if current_streak >= req_months:
        latest_month_str = qualified_months[-1]
        streak_tag = f"loyalty_streak_{latest_month_str}"
        
        existing_txn = PointsTransaction.query.filter_by(
            buyer_id=buyer.id,
            source='loyalty_streak_bonus',
            invoice_number=streak_tag
        ).first()
        
        if not existing_txn:
            buyer.total_points += bonus_pts
            db.session.add(PointsTransaction(
                buyer_id=buyer.id,
                points=bonus_pts,
                transaction_type='credit',
                source='loyalty_streak_bonus',
                invoice_number=streak_tag,
                rule_applied=f"BR-05 (Consecutive {req_months} Months Purchase Loyalty Bonus +{bonus_pts} pts)"[:100],
                created_at=datetime.utcnow()
            ))
            db.session.commit()
            return True, bonus_pts

    return False, 0

# ==================== CONFIGURATION & VERSIONING APIS ====================

@api.route('/admin/config', methods=['GET'])
@jwt_required()
def admin_get_config():
    current_user = get_jwt()
    if current_user['role'] != 'admin':
        return jsonify({'message': 'Unauthorized'}), 403
        
    config = Configuration.query.order_by(Configuration.version.desc()).first()
    if not config:
        # Fallback to seed default configuration
        config = Configuration(version=1)
        db.session.add(config)
        db.session.commit()
        
    inv_pct = getattr(config, 'invoice_reward_percentage', 0.50)
    if inv_pct is None:
        inv_pct = 0.50
        
    return jsonify({
        'version': config.version,
        'invoice_reward_percentage': inv_pct,
        'credit_period': config.credit_period,
        'forfeiture_cutoff': config.forfeiture_cutoff,
        'high_spend_threshold': config.high_spend_threshold,
        'high_spend_bonus': config.high_spend_bonus,
        'loyalty_consecutive_months': getattr(config, 'loyalty_consecutive_months', 3) or 3,
        'loyalty_min_monthly_purchase': getattr(config, 'loyalty_min_monthly_purchase', 200000.0) or 200000.0,
        'loyalty_bonus': config.loyalty_bonus if config.loyalty_bonus is not None else 10000,
        'regular_bonus': config.regular_bonus,
        'special_bonus': config.special_bonus,
        'old_stock_bonus': config.old_stock_bonus,
        'referral_min_value': config.referral_min_value,
        'referral_rate': config.referral_rate,
        'double_products': config.double_products,
        'shop_onboard_bonus': config.shop_onboard_bonus,
        'created_at': config.created_at.isoformat()
    }), 200

@api.route('/admin/config/versions', methods=['GET'])
@jwt_required()
def admin_get_config_versions():
    current_user = get_jwt()
    if current_user['role'] != 'admin':
        return jsonify({'message': 'Unauthorized'}), 403
        
    configs = Configuration.query.order_by(Configuration.version.desc()).all()
    
    result = []
    for c in configs:
        inv_pct = getattr(c, 'invoice_reward_percentage', 0.50)
        if inv_pct is None:
            inv_pct = 0.50
        result.append({
            'version': c.version,
            'invoice_reward_percentage': inv_pct,
            'credit_period': c.credit_period,
            'forfeiture_cutoff': c.forfeiture_cutoff,
            'high_spend_threshold': c.high_spend_threshold,
            'high_spend_bonus': c.high_spend_bonus,
            'loyalty_consecutive_months': getattr(c, 'loyalty_consecutive_months', 3) or 3,
            'loyalty_min_monthly_purchase': getattr(c, 'loyalty_min_monthly_purchase', 200000.0) or 200000.0,
            'loyalty_bonus': c.loyalty_bonus if c.loyalty_bonus is not None else 10000,
            'regular_bonus': c.regular_bonus,
            'special_bonus': c.special_bonus,
            'old_stock_bonus': c.old_stock_bonus,
            'referral_min_value': c.referral_min_value,
            'referral_rate': c.referral_rate,
            'double_products': c.double_products,
            'shop_onboard_bonus': c.shop_onboard_bonus,
            'created_at': c.created_at.isoformat()
        })
        
    return jsonify(result), 200

@api.route('/admin/config', methods=['POST'])
@jwt_required()
def admin_post_config():
    current_user = get_jwt()
    if current_user['role'] != 'admin':
        return jsonify({'message': 'Unauthorized'}), 403
        
    data = request.get_json()
    
    # Retrieve latest config
    latest = Configuration.query.order_by(Configuration.version.desc()).first()
    next_version = (latest.version + 1) if latest else 1
    
    new_config = Configuration(
        version=next_version,
        invoice_reward_percentage=float(data.get('invoice_reward_percentage', 0.50)),
        credit_period=int(data.get('credit_period', 7)),
        forfeiture_cutoff=int(data.get('forfeiture_cutoff', 30)),
        high_spend_threshold=float(data.get('high_spend_threshold', 200000.0)),
        high_spend_bonus=int(data.get('high_spend_bonus', 500)),
        loyalty_consecutive_months=int(data.get('loyalty_consecutive_months', 3)),
        loyalty_min_monthly_purchase=float(data.get('loyalty_min_monthly_purchase', 200000.0)),
        loyalty_bonus=int(data.get('loyalty_bonus', 10000)),
        regular_bonus=int(data.get('regular_bonus', 150)),
        special_bonus=int(data.get('special_bonus', 300)),
        old_stock_bonus=int(data.get('old_stock_bonus', 500)),
        referral_min_value=float(data.get('referral_min_value', 0.0)),
        referral_rate=float(data.get('referral_rate', 0.01)),
        double_products=data.get('double_products', 'Product X, Product Y'),
        shop_onboard_bonus=int(data.get('shop_onboard_bonus', 1000))
    )
    db.session.add(new_config)
    db.session.commit()
    
    return jsonify({'message': f'Settings updated successfully. Saved version {next_version}.', 'version': next_version}), 200

# ==================== PRODUCT MANAGEMENT APIS ====================

@api.route('/admin/products', methods=['GET'])
@jwt_required()
def admin_get_products():
    current_user = get_jwt()
    if current_user['role'] != 'admin':
        return jsonify({'message': 'Unauthorized'}), 403
        
    from datetime import datetime
    current_month = datetime.utcnow().strftime('%Y-%m')

    products = Product.query.order_by(Product.name).all()
    
    # Fetch all invoices for the current month in a single query
    invoices = Invoice.query.filter(
        func.to_char(Invoice.created_at, 'YYYY-MM') == current_month
    ).all()
    
    # Pre-lower products lists for case-insensitive containment check
    invoice_products = [inv.products.lower() if inv.products else "" for inv in invoices]
    
    result = []
    for p in products:
        p_name_lower = p.name.lower()
        invoices_this_month = sum(1 for inv_p in invoice_products if p_name_lower in inv_p)
        
        result.append({
            'id': p.id,
            'name': p.name,
            'tag': p.tag,
            'bonus_points': p.bonus_points,
            'category': p.category or 'Default',
            'brand': p.brand or 'Default',
            'sales_rate': p.sales_rate,
            'invoices_this_month': invoices_this_month
        })
    
    return jsonify(result), 200

@api.route('/admin/products', methods=['POST'])
@jwt_required()
def admin_post_product():
    current_user = get_jwt()
    if current_user['role'] != 'admin':
        return jsonify({'message': 'Unauthorized'}), 403
        
    data = request.get_json()
    name = data.get('name', '').strip()
    tag = data.get('tag', 'normal').strip().lower()
    bonus_points = int(data.get('bonus_points', 0) or 0)
    category = data.get('category', '').strip()
    brand = data.get('brand', '').strip()
    
    if not name:
        return jsonify({'message': 'Product name is required'}), 400
        
    product = Product.query.filter_by(name=name).first()
    if product:
        product.tag = tag
        product.bonus_points = bonus_points
        if category:
            product.category = category
        if brand:
            product.brand = brand
    else:
        product = Product(name=name, tag=tag, bonus_points=bonus_points, category=category, brand=brand)
        db.session.add(product)
        
    db.session.commit()
    return jsonify({'message': f'Product {name} updated successfully'}), 200

@api.route('/admin/products/<int:id>', methods=['DELETE'])
@jwt_required()
def admin_delete_product(id):
    current_user = get_jwt()
    if current_user['role'] != 'admin':
        return jsonify({'message': 'Unauthorized'}), 403
        
    product = Product.query.get(id)
    if not product:
        return jsonify({'message': 'Product not found'}), 404
        
    db.session.delete(product)
    db.session.commit()
    return jsonify({'message': 'Product deleted successfully'}), 200

# ==================== AUDIT LOG API ====================

@api.route('/admin/audit-logs', methods=['GET'])
@jwt_required()
def admin_audit_logs():
    current_user = get_jwt()
    if current_user['role'] != 'admin':
        return jsonify({'message': 'Unauthorized'}), 403
        
    search = request.args.get('search', '').strip()
    source_filter = request.args.get('source', '').strip()
    
    query = db.session.query(PointsTransaction, User).join(User, PointsTransaction.buyer_id == User.id)
    
    if source_filter:
        query = query.filter(PointsTransaction.source == source_filter)
        
    if search:
        search_pattern = f"%{search}%"
        query = query.filter(
            (User.name.like(search_pattern)) | 
            (User.email.like(search_pattern)) | 
            (PointsTransaction.invoice_number.like(search_pattern)) |
            (PointsTransaction.rule_applied.like(search_pattern))
        )
        
    txns = query.order_by(PointsTransaction.created_at.desc()).all()
    
    result = []
    for t in txns:
        txn = t.PointsTransaction
        user = t.User
        # Resolve config_version from the linked invoice if available
        config_ver = None
        if txn.invoice_number:
            inv = Invoice.query.filter_by(invoice_number=txn.invoice_number).first()
            if inv:
                config_ver = inv.config_version
        result.append({
            'id': txn.id,
            'buyer_name': user.name,
            'buyer_email': user.email,
            'buyer_role': user.role,
            'points': txn.points,
            'type': txn.transaction_type,
            'source': txn.source,
            'invoice_number': txn.invoice_number,
            'rule_applied': txn.rule_applied,
            'config_version': config_ver,
            'created_at': txn.created_at.isoformat()
        })
    
    return jsonify(result), 200

# ==================== DETAILED REP INVOICES API ====================

@api.route('/admin/reps/<int:rep_id>/invoices', methods=['GET'])
@jwt_required()
def admin_rep_invoices(rep_id):
    current_user = get_jwt()
    if current_user['role'] != 'admin':
        return jsonify({'message': 'Unauthorized'}), 403
        
    invoices = Invoice.query.filter_by(rep_id=rep_id).order_by(Invoice.created_at.desc()).all()
    
    result = []
    for inv in invoices:
        buyer = User.query.get(inv.buyer_id)
        
        config = Configuration.query.filter_by(version=inv.config_version).first()
        if not config:
            config = Configuration.query.order_by(Configuration.version.desc()).first()
            
        min_value = config.referral_min_value if config else 0.0
        qualifies = inv.amount >= min_value
        
        result.append({
            'id': inv.id,
            'invoice_number': inv.invoice_number,
            'buyer_name': buyer.business_name or buyer.name if buyer else 'N/A',
            'amount': inv.amount,
            'status': inv.status,
            'created_at': inv.created_at.isoformat(),
            'paid_at': inv.paid_at.isoformat() if inv.paid_at else None,
            'points_rep': inv.points_rep,
            'qualifies': qualifies,
            'min_value_required': min_value
        })
        
    return jsonify(result), 200

# ==================== ADVANCED REPORTS ANALYTICS API ====================

@api.route('/admin/reports/analytics', methods=['GET'])
@jwt_required()
def admin_reports_analytics():
    current_user = get_jwt()
    if current_user['role'] != 'admin':
        return jsonify({'message': 'Unauthorized'}), 403
        
    # 1. Financial liability
    total_unredeemed = db.session.query(func.sum(User.total_points)).filter(User.role != 'admin').scalar() or 0
    financial_liability = float(total_unredeemed) # 1 point = ₹1
    
    # Determine SQL function based on database engine dialect
    if db.engine.name == 'postgresql':
        month_func = func.to_char(PointsTransaction.created_at, 'YYYY-MM')
        invoice_month_func = func.to_char(Invoice.created_at, 'YYYY-MM')
    else:
        month_func = func.strftime('%Y-%m', PointsTransaction.created_at)
        invoice_month_func = func.strftime('%Y-%m', Invoice.created_at)

    # 2. Points issued vs redeemed (monthly chart data)
    issued_txns = db.session.query(
        month_func.label('month'),
        func.sum(PointsTransaction.points)
    ).filter(PointsTransaction.transaction_type == 'credit').group_by('month').all()
    
    redeemed_txns = db.session.query(
        month_func.label('month'),
        func.sum(PointsTransaction.points)
    ).filter(PointsTransaction.transaction_type == 'debit').group_by('month').all()
    
    issued_map = {item[0]: int(item[1]) for item in issued_txns if item[0]}
    redeemed_map = {item[0]: int(item[1]) for item in redeemed_txns if item[0]}
    
    all_months = sorted(list(set(issued_map.keys()) | set(redeemed_map.keys())))
    chart_data = []
    for m in all_months:
        chart_data.append({
            'name': m,
            'issued': issued_map.get(m, 0),
            'redeemed': redeemed_map.get(m, 0)
        })
        
    if not chart_data:
        chart_data = [
            {'name': '2026-05', 'issued': 4000, 'redeemed': 1200},
            {'name': '2026-06', 'issued': 6500, 'redeemed': 2500},
            {'name': '2026-07', 'issued': 12000, 'redeemed': 4800}
        ]
        
    # 3. Top Buyers
    top_buyers = []
    buyers = User.query.filter_by(role='buyer').all()
    for b in buyers:
        top_buyers.append({
            'id': b.id,
            'name': b.business_name or b.name,
            'total_points': b.total_points,
            'tier': 'Platinum' if b.total_points >= 10000 else 'Gold' if b.total_points >= 5000 else 'Silver'
        })
    top_buyers = sorted(top_buyers, key=lambda x: x['total_points'], reverse=True)[:5]
    
    # 4. Approaching spend threshold (within current calendar month)
    latest_config = Configuration.query.order_by(Configuration.version.desc()).first()
    threshold = latest_config.high_spend_threshold if latest_config else 200000.0
    
    current_month_str = datetime.utcnow().strftime('%Y-%m')
    buyer_monthly_spend = db.session.query(
        Invoice.buyer_id,
        func.sum(Invoice.amount)
    ).filter(
        Invoice.status == 'paid',
        invoice_month_func == current_month_str
    ).group_by(Invoice.buyer_id).all()
    
    approaching_customers = []
    for buyer_id, total_spend in buyer_monthly_spend:
        buyer = User.query.get(buyer_id)
        if buyer:
            pct = min(100.0, float(total_spend) / threshold * 100.0)
            # Threshold approaching lists (e.g. above 50% of threshold limit)
            approaching_customers.append({
                'id': buyer.id,
                'name': buyer.business_name or buyer.name,
                'email': buyer.email,
                'monthly_spend': float(total_spend),
                'threshold': threshold,
                'percentage': pct
            })
    approaching_customers = sorted(approaching_customers, key=lambda x: x['monthly_spend'], reverse=True)
    
    # 5. Marketing rep monthly payouts summary
    rep_payouts_raw = db.session.query(
        PointsTransaction.buyer_id,
        month_func.label('month'),
        func.sum(PointsTransaction.points)
    ).filter(
        PointsTransaction.source == 'sales_commission'
    ).group_by(PointsTransaction.buyer_id, 'month').all()
    
    payout_summaries = []
    for rep_id, month, total_points in rep_payouts_raw:
        rep = User.query.get(rep_id)
        if rep:
            payout_summaries.append({
                'rep_id': rep.id,
                'rep_name': rep.name,
                'rep_email': rep.email,
                'month': month,
                'commission_points': int(total_points),
                'amount_rupees': float(total_points)
            })
    payout_summaries = sorted(payout_summaries, key=lambda x: (x['month'], x['commission_points']), reverse=True)
    
    return jsonify({
        'liability': {
            'total_unredeemed_points': total_unredeemed,
            'financial_liability': financial_liability
        },
        'chart_data': chart_data,
        'top_buyers': top_buyers,
        'approaching_customers': approaching_customers,
        'payout_summaries': payout_summaries
    }), 200
